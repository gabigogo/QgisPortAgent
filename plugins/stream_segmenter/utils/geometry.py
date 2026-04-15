"""Pure-Python constants and geometry helpers for the stream_segmenter plugin.

All symbols in this module depend only on ``shapely`` and the Python standard
library — no QGIS bindings required.  This separation allows helpers and
constants to be unit-tested without a running QGIS environment.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional, Set, Tuple

# ── Unit conversion constants ─────────────────────────────────────────────────

MILES_TO_METRES: float = 1609.344
KM_TO_METRES: float = 1000.0

# ── Output field names ────────────────────────────────────────────────────────

FLD_SEG_NUM: str = "seg_num"
FLD_MILE_RANGE: str = "mile_range"
FLD_STREAM_SEG_ID: str = "stream_seg_id"
FLD_LENGTH_MI: str = "length_mi"
NEW_FIELD_NAMES: frozenset = frozenset({FLD_SEG_NUM, FLD_MILE_RANGE, FLD_STREAM_SEG_ID, FLD_LENGTH_MI})


def sanitize_name(raw: str) -> str:
    """Strip surrounding whitespace and replace interior spaces with underscores.

    Args:
        raw (str): Unsanitized stream name string.

    Returns:
        str: Sanitized name safe for use in field values and identifiers.

    Examples:
        >>> sanitize_name("  Bear Creek  ")
        'Bear_Creek'
        >>> sanitize_name("")
        ''
    """
    return raw.strip().replace(" ", "_")


def fmt_distance(v: float) -> str:
    """Format a distance value as an integer string when whole, else 1 decimal.

    Args:
        v (float): Numeric value to format (typically a mile or km boundary).

    Returns:
        str: ``"0"``, ``"5"``, ``"10.5"``, etc.

    Examples:
        >>> fmt_distance(0.0)
        '0'
        >>> fmt_distance(5.0)
        '5'
        >>> fmt_distance(2.5)
        '2.5'
    """
    return str(int(v)) if v == int(v) else f"{v:.1f}"


def to_oriented_line(wkb_bytes: bytes) -> Tuple[Optional[Any], Optional[str]]:
    """Convert WKB bytes to a shapely LineString oriented from the downstream end.

    "Oriented from downstream" means the **first coordinate** of the returned
    LineString is the downstream endpoint of the original feature.  The
    downstream endpoint is defined as:

    - **LineString**: last vertex of the geometry.
    - **MultiLineString**: last vertex of the last part.

    For MultiLineString inputs, ``shapely.ops.linemerge`` resolves parts into
    a single connected chain before orientation is applied.

    Args:
        wkb_bytes (bytes): Well-Known Binary representation of the line geometry.

    Returns:
        Tuple[Optional[LineString], Optional[str]]:
            - ``(oriented_line, None)`` on success.
            - ``(None, error_message)`` on failure.

    Assumptions:
        - WKB represents a LineString or MultiLineString geometry.
        - ``shapely >= 2.0`` is available.
        - MultiLineString parts form a topologically connected chain.
        - Coordinate tolerance for endpoint matching: 1e-6.

    Methodology:
        1. Deserialise WKB via ``shapely.wkb.loads``.
        2. For MultiLineString: record downstream coord (last coord of last
           part), run linemerge, orient the result so first point = downstream.
        3. For LineString: reverse coordinate sequence so last vertex becomes
           first.

    Examples:
        >>> from shapely.geometry import LineString
        >>> line = LineString([(0, 0), (5, 0), (10, 0)])
        >>> oriented, err = to_oriented_line(bytes(line.wkb))
        >>> err is None
        True
        >>> list(oriented.coords)[0]
        (10.0, 0.0)
    """
    from shapely import wkb as swkb
    from shapely.geometry import LineString
    from shapely.ops import linemerge

    try:
        shapely_geom = swkb.loads(wkb_bytes)
    except Exception as exc:
        return None, f"WKB conversion error: {exc}"

    if shapely_geom is None or shapely_geom.is_empty:
        return None, "empty geometry after WKB conversion"

    # ── MultiLineString ────────────────────────────────────────────────────
    if hasattr(shapely_geom, "geoms"):
        parts = list(shapely_geom.geoms)
        downstream_coord = parts[-1].coords[-1]

        merged = linemerge(shapely_geom)

        if hasattr(merged, "geoms"):
            return None, (
                "MultiLineString parts are disconnected and cannot be merged "
                "into a single chain — verify the feature's geometry topology"
            )

        tol = 1e-6
        mc0 = merged.coords[0]
        mc_last = merged.coords[-1]

        if (
            abs(mc_last[0] - downstream_coord[0]) < tol
            and abs(mc_last[1] - downstream_coord[1]) < tol
        ):
            # Last coord = downstream. Reverse so first = downstream.
            return LineString(list(merged.coords)[::-1]), None

        if (
            abs(mc0[0] - downstream_coord[0]) < tol
            and abs(mc0[1] - downstream_coord[1]) < tol
        ):
            # First coord already = downstream.
            return merged, None

        return None, (
            "Downstream endpoint could not be matched after MultiLineString "
            "merge (tolerance 1e-6). Verify that the last part's last vertex "
            "is the channel mouth."
        )

    # ── Single LineString ──────────────────────────────────────────────────
    # Last coord = downstream. Reverse so first coord = downstream.
    return LineString(list(shapely_geom.coords)[::-1]), None


def cut_fraction(
    line: Any,
    start_frac: float,
    end_frac: float,
) -> Optional[Any]:
    """Cut a shapely LineString between two normalised fractions [0.0, 1.0].

    Uses ``shapely.ops.substring`` with ``normalized=True`` so positions are
    expressed as fractions of the total Euclidean arc length.

    Args:
        line: Shapely ``LineString`` to cut.  Must be single-part.
        start_frac (float): Start fraction [0, 1].
        end_frac (float): End fraction [0, 1].  Must be >= *start_frac*.

    Returns:
        Optional[LineString]: The cut segment, or ``None`` for empty /
        degenerate (< 2 vertices, zero-length) results.

    Assumptions:
        - *line* is a valid, non-empty single-part ``LineString``.
        - 0.0 <= start_frac <= end_frac <= 1.0.
        - ``shapely >= 2.0`` is available.

    Methodology:
        1. Call ``shapely.ops.substring(line, start_frac, end_frac, normalized=True)``.
        2. Resolve any degenerate MultiLineString result by choosing the
           longest non-empty part.
        3. Discard results with < 2 vertices or zero length.

    Examples:
        >>> from shapely.geometry import LineString
        >>> line = LineString([(0, 0), (10, 0)])
        >>> seg = cut_fraction(line, 0.0, 0.5)
        >>> list(seg.coords)
        [(0.0, 0.0), (5.0, 0.0)]
    """
    from shapely.ops import substring

    try:
        seg = substring(line, start_frac, end_frac, normalized=True)
    except Exception:
        return None

    if seg is None or seg.is_empty:
        return None

    # Resolve occasional MultiLineString from degenerate cuts.
    if hasattr(seg, "geoms"):
        parts = [g for g in seg.geoms if not g.is_empty]
        if not parts:
            return None
        seg = max(parts, key=lambda g: g.length)

    # Discard degenerate zero-length or single-vertex segments.
    if len(list(seg.coords)) < 2 or seg.length == 0.0:
        return None

    return seg


def parse_range_start(mile_range: str) -> Optional[float]:
    """Extract the start-distance value from a ``mile_range`` label.

    The ``mile_range`` field produced by ``StreamSegmenterAlgorithm`` has the
    format ``"{start}-{end} {unit}"``, e.g. ``"0-1 mi"``, ``"4.5-5 km"``.
    This function returns the numeric start value.

    Args:
        mile_range (str): Range label string as written to the output layer.

    Returns:
        Optional[float]: The start distance (in whatever user units the label
        was written with), or ``None`` if the string cannot be parsed.

    Examples:
        >>> parse_range_start("0-1 mi")
        0.0
        >>> parse_range_start("4-5 mi")
        4.0
        >>> parse_range_start("4.5-5 km")
        4.5
        >>> parse_range_start("bad") is None
        True
    """
    try:
        # "0-1 mi"  →  "0-1"  →  "0"  →  0.0
        # "4.5-5 km" →  "4.5-5"  →  "4.5"  →  4.5
        range_part = mile_range.strip().split()[0]
        start_str = range_part.split("-")[0]
        return float(start_str)
    except (IndexError, ValueError, AttributeError):
        return None


# ── File-system helpers ───────────────────────────────────────────────────────

#: OGR-readable line vector extensions used when no explicit filter is supplied.
DEFAULT_VECTOR_EXTENSIONS: frozenset = frozenset(
    [".gpkg", ".shp", ".geojson", ".json", ".fgb", ".gml", ".kml"]
)


def collect_files(folder: "Path", filter_str: str) -> "List[Path]":
    """Return all files in *folder* matching the glob patterns in *filter_str*.

    Non-recursive: only top-level files are scanned.

    Args:
        folder (Path): Directory to scan.
        filter_str (str): Space-separated glob patterns, e.g. ``"*.gpkg *.shp"``.
            An empty string or ``"*"`` returns all files whose suffix is in
            ``DEFAULT_VECTOR_EXTENSIONS``.

    Returns:
        List[Path]: Sorted, deduplicated list of matching file paths.

    Examples:
        >>> from pathlib import Path
        >>> collect_files(Path("/data/streams"), "*.gpkg *.shp")
        [PosixPath('/data/streams/creek.gpkg'), PosixPath('/data/streams/river.shp')]
    """
    from pathlib import Path
    from typing import List

    patterns = [p.strip() for p in filter_str.split() if p.strip()]
    if not patterns or patterns == ["*"]:
        return sorted(
            f for f in folder.iterdir()
            if f.is_file() and f.suffix.lower() in DEFAULT_VECTOR_EXTENSIONS
        )
    seen: set = set()
    results: "List[Path]" = []
    for pat in patterns:
        for match in sorted(folder.glob(pat)):
            if match.is_file() and match not in seen:
                seen.add(match)
                results.append(match)
    return results


def load_reference_values(
    table_path: str,
    column_selector: str,
    has_header: bool = True,
) -> Set[str]:
    """Load unique, non-empty values from one column in a CSV or XLSX table.

    Args:
        table_path (str): Path to a CSV or XLSX reference table.
        column_selector (str): Column name (when *has_header* is True) or
            zero-based column index as a string (e.g. ``"0"``).
        has_header (bool): Whether the table has a header row.

    Returns:
        Set[str]: Normalized unique values from the selected column.

    Raises:
        ValueError: If the file extension is unsupported, the column cannot be
            resolved, or the table cannot be parsed.
    """
    path = Path(table_path)
    if not path.exists() or not path.is_file():
        raise ValueError(f"Reference file does not exist: {table_path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        return _load_reference_values_csv(path, column_selector, has_header)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_reference_values_xlsx(path, column_selector, has_header)

    raise ValueError(
        "Unsupported reference file format. Use CSV or XLSX."
    )


def _load_reference_values_csv(
    path: Path,
    column_selector: str,
    has_header: bool,
) -> Set[str]:
    import csv

    values: Set[str] = set()

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
    except Exception as exc:
        raise ValueError(f"Failed reading CSV reference file: {exc}") from exc

    if not rows:
        return values

    col_idx = _resolve_column_index(rows[0], column_selector, has_header)
    start_idx = 1 if has_header else 0

    for row in rows[start_idx:]:
        if col_idx >= len(row):
            continue
        raw = row[col_idx]
        if raw is None:
            continue
        norm = str(raw).strip().lower()
        if norm:
            values.add(norm)

    return values


def _load_reference_values_xlsx(
    path: Path,
    column_selector: str,
    has_header: bool,
) -> Set[str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError(
            "Excel support requires openpyxl to be installed."
        ) from exc

    values: Set[str] = set()

    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed reading Excel reference file: {exc}") from exc

    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return values

    header = ["" if cell is None else str(cell).strip() for cell in rows[0]]
    col_idx = _resolve_column_index(header, column_selector, has_header)
    start_idx = 1 if has_header else 0

    for row in rows[start_idx:]:
        if row is None or col_idx >= len(row):
            continue
        raw = row[col_idx]
        if raw is None:
            continue
        norm = str(raw).strip()
        if norm:
            values.add(norm)

    return values


def _resolve_column_index(
    header: list,
    column_selector: str,
    has_header: bool,
) -> int:
    selector = (column_selector or "").strip()
    if selector == "":
        raise ValueError("Reference column cannot be empty.")

    if selector.isdigit():
        return int(selector)

    if not has_header:
        raise ValueError(
            "Non-numeric column selector requires a header row."
        )

    lowered = [str(v).strip().lower() for v in header]
    try:
        return lowered.index(selector.lower())
    except ValueError as exc:
        raise ValueError(
            f"Column '{column_selector}' was not found in the reference table."
        ) from exc


def matches_reference(
    candidate: Any,
    reference_values: Set[str],
    exact: bool,
) -> bool:
    """Return whether a candidate value matches the provided reference set.

    Args:
        candidate (Any): Feature attribute value to evaluate.
        reference_values (Set[str]): Normalized reference values.
        exact (bool): True for exact-match mode, False for partial contains mode.

    Returns:
        bool: ``True`` if the candidate matches according to the selected mode.
    """
    if candidate is None or not reference_values:
        return False

    candidate_norm = str(candidate).strip().lower()
    if not candidate_norm:
        return False

    if exact:
        return candidate_norm in reference_values

    return any(ref in candidate_norm for ref in reference_values)
